#!/usr/bin/env python3
"""Backfill FAST/SPS sinter runs from the FCT HP D 250 log workbooks into Directus.

Reads every monthly tab of each `FAST Data/*.xlsx`, maps rows (see fast_mapping.py)
to `manufacturing_operations` (method MF, process_category 'sintering') plus the
inline `sintering_*` fields, creates the approved new `materials` and any
`Machine_Operators` it needs, and links `owner` to existing Directus users.

Idempotent: each run is keyed by a deterministic `source_run_uid`
(machine|date|time|batch); re-runs INSERT … ON CONFLICT DO NOTHING.

Usage:
    DATABASE_URL=postgres://d1:change_me@localhost:5432/d1_database \
        python scripts/import_fast_logs.py "FAST Data" [--dry-run]
"""
from __future__ import annotations

import argparse
import glob
import os
import re
import sys
import uuid
from collections import Counter
from datetime import date, datetime, time

import openpyxl
import psycopg2
import psycopg2.extras

import fast_mapping as fm

LEGACY_NOTE = "Imported from FAST/SPS log sheets (FCT HP D 250)."
_NS = uuid.uuid5(uuid.NAMESPACE_DNS, "d1-database.fast-log.v1")

# Normalised source header → canonical field key (strips the ° so encoding can't bite).
CANON_HEADERS = {
    "date": "date", "time": "time", "user": "user", "batch #": "batch",
    "recipe #": "recipe", "material": "material", "mass (g)": "mass",
    "mould diameter (mm)": "mould", "atmosphere": "atmosphere",
    "tc/pyro control": "tcpyro", "max force (kn)": "maxforce",
    "max temp (c)": "maxtemp", "voltage at max t (v)": "voltage",
    "power at max t (kw)": "power", "ptc top (c)": "ptctop",
    "ptc bot (c)": "ptcbot", "comments, failures, alarms": "comments",
}


def _hkey(h: str) -> str:
    return re.sub(r"\s+", " ", str(h).strip().lower().replace("°", ""))


def clean_str(v):
    s = str(v).strip() if v is not None else ""
    return s if s and s.lower() not in ("none", "n/a", "na", "") else None


def clean_float(v):
    if v is None:
        return None
    try:
        f = float(v)
        return None if f == 0.0 else f
    except (ValueError, TypeError):
        return None


def clean_code(v):
    """Code-like fields (batch/recipe): render whole-number floats as ints."""
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    s = clean_str(v)
    if s and re.fullmatch(r"\d+\.0", s):
        return s[:-2]
    return s


def fmt_date(v):
    if isinstance(v, datetime):
        v = v.date()
    if isinstance(v, date):
        return None if v.year < 1990 else v
    s = clean_str(v)
    if not s:
        return None
    for f in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s[:10], f).date()
        except ValueError:
            pass
    return None


def fmt_time(v):
    if isinstance(v, (time, datetime)):
        return v.strftime("%H:%M")
    s = clean_str(v)
    return s or ""


def read_runs(folder: str) -> list[dict]:
    """Return canonical-keyed run dicts across all workbooks/tabs."""
    runs = []
    for path in sorted(glob.glob(os.path.join(folder, "*.xlsx"))):
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for tab in wb.sheetnames:
            ws = wb[tab]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            colmap = {}  # column index → canonical key
            for i, h in enumerate(rows[0]):
                if h is None:
                    continue
                ck = CANON_HEADERS.get(_hkey(h))
                if ck:
                    colmap[i] = ck
            if "date" not in colmap.values():
                continue
            for r in rows[1:]:
                rec = {ck: (r[i] if i < len(r) else None) for i, ck in colmap.items()}
                if rec.get("date") is None:
                    continue
                if not any(clean_str(rec.get(c)) for c in ("batch", "recipe", "material", "mass")):
                    continue
                runs.append(rec)
    return runs


def build(run: dict) -> dict | None:
    d = fmt_date(run.get("date"))
    if d is None:
        return None
    t = fmt_time(run.get("time"))
    batch = clean_code(run.get("batch"))
    uid = fm.run_uid(fm.SOURCE_MACHINE, str(d), t, batch or "")
    operator_name, owner_email, raw_user = fm.parse_user(clean_str(run.get("user")))
    material_text = clean_str(run.get("material"))
    return {
        "source_run_uid": uid,
        "operation_id": str(uuid.uuid5(_NS, uid)),
        "operation_date": d,
        "operator_name": raw_user,           # raw User string preserved
        "operator_disp": operator_name,      # resolved → Machine_Operators
        "owner_email": owner_email,
        "material_code": fm.match_material(material_text),
        "sintering_recipe_number": clean_code(run.get("recipe")),
        "sintering_batch_number": batch,
        "sintering_mass_grams": clean_float(run.get("mass")),
        "sintering_mould_diameter_mm": clean_float(run.get("mould")),
        "sintering_atmosphere": clean_str(run.get("atmosphere")),
        "sintering_tc_pyro_control": clean_str(run.get("tcpyro")),
        "sintering_max_force_kn": clean_float(run.get("maxforce")),
        "sintering_max_temp_celsius": clean_float(run.get("maxtemp")),
        "sintering_voltage_at_max_t_v": clean_float(run.get("voltage")),
        "sintering_power_at_max_t_kw": clean_float(run.get("power")),
        "sintering_ptc_top_celsius": clean_float(run.get("ptctop")),
        "sintering_ptc_bot_celsius": clean_float(run.get("ptcbot")),
        "sintering_material_type_note": material_text,
        "outcome_notes": " ".join(filter(None, [LEGACY_NOTE, clean_str(run.get("comments"))])),
    }


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("folder", help="folder containing the FAST log xlsx files")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    runs = [b for b in (build(r) for r in read_runs(args.folder)) if b]
    # de-dupe within the import by source_run_uid (last wins)
    by_uid = {b["source_run_uid"]: b for b in runs}
    ops = list(by_uid.values())

    mats_needed = Counter(b["material_code"] for b in ops if b["material_code"])
    operators_needed = sorted({b["operator_disp"] for b in ops if b["operator_disp"]})
    owners_matched = sum(1 for b in ops if b["owner_email"])
    no_owner = sum(1 for b in ops if not b["owner_email"])
    no_material = sum(1 for b in ops if not b["material_code"])

    print(f"Parsed {len(runs)} run-rows → {len(ops)} unique operations")
    print(f"  material linked: {len(ops) - no_material}   free-text only: {no_material}")
    print(f"  owner matched to app user: {owners_matched}   no owner: {no_owner}")
    print(f"  distinct operators (Machine_Operators): {len(operators_needed)}")
    print(f"  new materials to create: {len(fm.NEW_MATERIALS)}")
    print("  top linked material codes:",
          ", ".join(f"{c}×{n}" for c, n in mats_needed.most_common(12)))
    print("  operators:", ", ".join(operators_needed[:25]), "…" if len(operators_needed) > 25 else "")

    # Sanity-flag implausible dates (almost always a source-sheet typo, e.g. "17/08/0225").
    # Imported faithfully, but surfaced so the maintainer can correct the log sheet / DB.
    this_year = date.today().year
    bad_dates = [b for b in ops if b["operation_date"] and not (2015 <= b["operation_date"].year <= this_year + 1)]
    if bad_dates:
        print(f"  WARNING: {len(bad_dates)} run(s) have an implausible date (likely a source typo) — review:")
        for b in bad_dates[:20]:
            print(f"     {b['operation_date']}  recipe={b['sintering_recipe_number']}  {b['sintering_material_type_note']}")

    if args.dry_run:
        print("\n[dry-run] no writes. Sample of 3 mapped operations:")
        for b in ops[:3]:
            print("   ", {k: b[k] for k in ("operation_date", "operator_disp", "owner_email",
                                            "material_code", "sintering_recipe_number",
                                            "sintering_max_temp_celsius", "sintering_mass_grams")})
        return

    db = os.environ.get("DATABASE_URL")
    if not db:
        sys.exit("ERROR: DATABASE_URL required for a real import")
    conn = psycopg2.connect(db)
    conn.autocommit = False
    cur = conn.cursor()

    # 1) create approved new materials (idempotent)
    psycopg2.extras.execute_values(
        cur,
        "INSERT INTO materials (material_id, alloy_code, common_name, notes) VALUES %s "
        "ON CONFLICT DO NOTHING",
        [(fm.material_id(code), code, name, "FAST log import") for code, name in fm.NEW_MATERIALS],
    )
    # 2) resolve every code → material_id (existing + new)
    cur.execute("SELECT alloy_code, material_id FROM materials")
    code_to_mat = {c: m for c, m in cur.fetchall()}

    # 3) ensure Machine_Operators for each distinct operator; map name → id
    cur.execute('SELECT "Name", id FROM "Machine_Operators"')
    op_to_id = {n: i for n, i in cur.fetchall()}
    for name in operators_needed:
        if name not in op_to_id:
            cur.execute('INSERT INTO "Machine_Operators" ("Name") VALUES (%s) RETURNING id', (name,))
            op_to_id[name] = cur.fetchone()[0]

    # 4) owner email → directus_users id
    cur.execute("SELECT email, id FROM directus_users")
    email_to_user = {e: i for e, i in cur.fetchall()}

    method_id = "4b13b4f4-7f7e-5356-b93a-937ab527386d"   # MF
    equipment_id = "27f468ae-5e5b-532d-bf33-e9cfc939b524"  # FCT HP D 250

    cols = [
        "operation_id", "method_id", "equipment_id", "process_category",
        "source_run_uid", "source_system", "operation_date",
        "operator", "operator_name", "owner", "material_id",
        "sintering_recipe_number", "sintering_batch_number", "sintering_mass_grams",
        "sintering_mould_diameter_mm", "sintering_atmosphere", "sintering_tc_pyro_control",
        "sintering_max_force_kn", "sintering_max_temp_celsius", "sintering_voltage_at_max_t_v",
        "sintering_power_at_max_t_kw", "sintering_ptc_top_celsius", "sintering_ptc_bot_celsius",
        "sintering_material_type_note", "outcome_notes",
    ]
    values = []
    for b in ops:
        values.append((
            b["operation_id"], method_id, equipment_id, "sintering",
            b["source_run_uid"], fm.SOURCE_SYSTEM, b["operation_date"],
            op_to_id.get(b["operator_disp"]), b["operator_name"],
            email_to_user.get(b["owner_email"]) if b["owner_email"] else None,
            code_to_mat.get(b["material_code"]) if b["material_code"] else None,
            b["sintering_recipe_number"], b["sintering_batch_number"], b["sintering_mass_grams"],
            b["sintering_mould_diameter_mm"], b["sintering_atmosphere"], b["sintering_tc_pyro_control"],
            b["sintering_max_force_kn"], b["sintering_max_temp_celsius"], b["sintering_voltage_at_max_t_v"],
            b["sintering_power_at_max_t_kw"], b["sintering_ptc_top_celsius"], b["sintering_ptc_bot_celsius"],
            b["sintering_material_type_note"], b["outcome_notes"],
        ))
    col_sql = ", ".join(cols)
    psycopg2.extras.execute_values(
        cur,
        f"INSERT INTO manufacturing_operations ({col_sql}) VALUES %s "
        f"ON CONFLICT (source_run_uid) WHERE source_run_uid IS NOT NULL DO NOTHING",
        values,
    )
    inserted = cur.rowcount
    conn.commit()
    print(f"\nInserted {inserted} new FAST operations "
          f"({len(ops) - inserted} already present / skipped).")
    cur.close()
    conn.close()


if __name__ == "__main__":
    main()
