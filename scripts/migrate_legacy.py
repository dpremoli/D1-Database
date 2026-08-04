#!/usr/bin/env python3
"""Phase 8 — Legacy data migration from Sample_Data.xlsx into D1 PostgreSQL.

Reads the AppSheet/Sheets export and loads every entity into the target schema
in FK-dependency order. Safe to re-run: existing rows are skipped via
ON CONFLICT DO NOTHING (materials use COALESCE to back-fill missing density).

Provenance: every imported row records 'Imported from legacy AppSheet export'
in its notes / outcome_notes field so the audit trail is clear.

Usage
-----
    pip install -r scripts/requirements.txt
    DATABASE_URL="postgres://d1:change_me@localhost:5432/d1_database" \\
        python3 scripts/migrate_legacy.py --xlsx /path/to/Sample_Data.xlsx

    # Dry-run (prints row counts only, no DB writes):
    python3 scripts/migrate_legacy.py --xlsx /path/to/Sample_Data.xlsx --dry-run
"""

from __future__ import annotations

import argparse
import json
import logging
import os
import re
import sys
import uuid
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl
import psycopg2
import psycopg2.extras

log = logging.getLogger("migrate_legacy")

# Stable UUID namespace — every re-run of this script produces the same UUIDs
# for the same AppSheet legacy IDs, making the migration safely idempotent.
_LEGACY_NS = uuid.uuid5(uuid.NAMESPACE_DNS, "d1-database.legacy-migration.v1")

LEGACY_NOTE = "Imported from legacy AppSheet/Sheets export (Phase 8 migration)."

_LAB_ADMIN_ROLE_ID = "10000001-0000-0000-0000-000000000001"
_LAB_MEMBER_ROLE_ID = "10000001-0000-0000-0000-000000000002"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def legacy_uuid(raw_id: Any) -> uuid.UUID:
    """Deterministic UUID from any AppSheet legacy ID (hex, float-str, etc.)."""
    return uuid.uuid5(_LEGACY_NS, str(raw_id).strip())


def clean_str(v: Any) -> str | None:
    s = str(v).strip() if v is not None else None
    return s if s and s.lower() not in ("none", "n/a", "na") else None


def clean_float(v: Any) -> float | None:
    if v is None:
        return None
    try:
        f = float(v)
        return None if f == 0.0 else f
    except (ValueError, TypeError):
        return None


def clean_date(v: Any) -> date | None:
    """Parse Excel datetime / date string; return None for garbage dates."""
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        d = v if isinstance(v, date) else v.date()
        # Excel zero epoch: 1899-12-30 == empty cell
        if d.year == 1899:
            return None
        return d
    s = str(v).strip()
    # "1899-12-30: 00:00:00" pattern = empty
    if s.startswith("1899"):
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except ValueError:
            pass
    return None


def clean_bool(v: Any) -> bool:
    if isinstance(v, bool):
        return v
    if isinstance(v, str):
        return v.strip().lower() in ("true", "yes", "1")
    return bool(v) if v is not None else False


def sheet_rows(
    wb: openpyxl.Workbook, name: str, required: bool = True
) -> tuple[list[str], list[dict]]:
    """Return (headers, list_of_dicts) for a sheet, skipping blank rows.

    If required=False and the sheet is absent, returns ([], []) instead of
    raising KeyError.
    """
    if name not in wb.sheetnames:
        if not required:
            log.warning("Sheet %r not found in workbook — skipping", name)
            return [], []
        raise KeyError(f"Sheet {name!r} not found in workbook")
    ws = wb[name]
    raw = list(ws.iter_rows(values_only=True))
    headers = [str(h) for h in raw[0] if h is not None]
    rows = []
    for r in raw[1:]:
        if not any(v is not None for v in r):
            continue
        rows.append(dict(zip(headers, r[: len(headers)])))
    return headers, rows


def extract_tool_uuid(tool_field: str | None) -> str | None:
    """Extract hex legacy ID from '(072192a6) DDJNL2525X15JETI' format."""
    if not tool_field:
        return None
    m = re.match(r"\(([0-9a-fA-F]+)\)", str(tool_field).strip())
    return m.group(1) if m else None


# ---------------------------------------------------------------------------
# Loaders — one function per target table
# ---------------------------------------------------------------------------


def load_alloying_elements(cur, rows: list[dict], dry: bool) -> int:
    data = []
    for r in rows:
        sym = clean_str(r.get("Symbol"))
        name = clean_str(r.get("Name"))
        num = r.get("Atomic Number")
        if not (sym and name and num):
            continue
        data.append(
            (
                sym,
                name,
                int(float(num)),
                clean_float(r.get("Atomic Weight")),
                clean_float(r.get("Density (g/cm3)")),
                clean_float(r.get("Melting Point (K)")),
                clean_float(r.get("Boiling Point (K)")),
                clean_float(r.get("Electronegativity")),
                clean_float(r.get("Atomic Radius (pm)")),
            )
        )
    if not dry:
        psycopg2.extras.execute_values(
            cur,
            """INSERT INTO alloying_elements
                 (symbol, element_name, atomic_number,
                  atomic_weight, density_g_per_cm3, melting_point_k,
                  boiling_point_k, electronegativity, atomic_radius_pm)
               VALUES %s
               ON CONFLICT (symbol) DO UPDATE SET
                 atomic_weight     = COALESCE(EXCLUDED.atomic_weight, alloying_elements.atomic_weight),
                 density_g_per_cm3 = COALESCE(EXCLUDED.density_g_per_cm3, alloying_elements.density_g_per_cm3),
                 melting_point_k   = COALESCE(EXCLUDED.melting_point_k, alloying_elements.melting_point_k),
                 boiling_point_k   = COALESCE(EXCLUDED.boiling_point_k, alloying_elements.boiling_point_k),
                 electronegativity = COALESCE(EXCLUDED.electronegativity, alloying_elements.electronegativity),
                 atomic_radius_pm  = COALESCE(EXCLUDED.atomic_radius_pm, alloying_elements.atomic_radius_pm)""",
            data,
        )
    return len(data)


def load_materials(cur, rows: list[dict], dry: bool) -> int:
    data = []
    for r in rows:
        code = clean_str(r.get("Code"))
        name = clean_str(r.get("Alloy"))
        if not (code and name):
            continue
        density = clean_float(r.get("Density"))
        export_ctrl = clean_bool(r.get("Export Controlled?"))
        datasheet = clean_str(r.get("Link"))
        overview = clean_str(r.get("Overview"))
        notes = f"{LEGACY_NOTE} Overview: {overview}" if overview else LEGACY_NOTE
        data.append(
            (
                str(legacy_uuid(code)),  # stable material_id
                code,
                name,
                density,
                export_ctrl,
                datasheet,
                notes[:2000] if notes else LEGACY_NOTE,
            )
        )
    # Deduplicate by alloy_code (index 1) — keep first occurrence
    seen: set[str] = set()
    deduped = []
    for row in data:
        if row[1] not in seen:
            seen.add(row[1])
            deduped.append(row)
    data = deduped
    if not dry:
        psycopg2.extras.execute_values(
            cur,
            """INSERT INTO materials
                 (material_id, alloy_code, common_name, density_g_per_cm3,
                  export_controlled, datasheet_url, notes)
               VALUES %s
               ON CONFLICT (alloy_code) DO UPDATE SET
                 density_g_per_cm3 = COALESCE(materials.density_g_per_cm3,
                                               EXCLUDED.density_g_per_cm3),
                 datasheet_url = COALESCE(materials.datasheet_url,
                                          EXCLUDED.datasheet_url)""",
            data,
        )
    return len(data)


def load_material_elements(cur, rows: list[dict], dry: bool) -> int:
    """Populate material_alloying_elements M2M junction from Alloy Codes sheet.

    The 'Alloying Elements' column in AppSheet is a comma-separated EnumList of
    element Symbols (e.g. 'Ti, Al, V').  We resolve each symbol against the
    already-loaded alloying_elements table.
    """
    if not dry:
        cur.execute("SELECT symbol FROM alloying_elements")
        known_symbols = {r[0] for r in cur.fetchall()}
        cur.execute("SELECT alloy_code, material_id FROM materials")
        code_to_uuid = {r[0]: r[1] for r in cur.fetchall()}
    else:
        known_symbols = set()
        code_to_uuid = {}

    def _extract_symbol(token: str, known: set[str]) -> str | None:
        """Extract a clean element symbol from an AppSheet composition token.

        AppSheet sometimes stores "6% Al", "bal Ni", "16–18% Cr", "Ni 58.0%",
        "C 0.43-0.5" etc.  We try every space-split word against the known set.
        Also handles multi-character symbols like "Mo", "Hf", "Ti" etc.
        """
        token = token.strip()
        if token.lower() in ("n/a", "na", "others", "other", "balance", "bal"):
            return None
        # Direct match first
        if token in known:
            return token
        # Try each word in the token
        for word in re.split(r"[\s%–\-/:.]+", token):
            word = word.strip()
            if word in known:
                return word
            # Capitalise first letter (AppSheet may store lowercase)
            cap = word.capitalize()
            if cap in known:
                return cap
        return None

    data = []
    for r in rows:
        code = clean_str(r.get("Code"))
        if not code:
            continue
        mat_id = code_to_uuid.get(code) or str(legacy_uuid(code))
        raw = r.get("Alloying Elements")
        if not raw:
            continue
        seen_syms: set[str] = set()
        for token in str(raw).split(","):
            sym = (
                _extract_symbol(token.strip(), known_symbols)
                if known_symbols
                else token.strip()
            )
            if not sym:
                if known_symbols and token.strip().lower() not in (
                    "n/a",
                    "na",
                    "others",
                    "",
                ):
                    log.debug(
                        "material_elements: no match for %r, skipping", token.strip()
                    )
                continue
            if sym not in seen_syms:
                seen_syms.add(sym)
                data.append((mat_id, sym))

    if not dry and data:
        psycopg2.extras.execute_values(
            cur,
            """INSERT INTO material_alloying_elements (material_id, symbol)
               VALUES %s ON CONFLICT DO NOTHING""",
            data,
        )
    return len(data)


def load_equipment(cur, rows: list[dict], dry: bool) -> int:
    data = []
    for r in rows:
        name = clean_str(r.get("Machine"))
        if not name:
            continue
        code = re.sub(r"[^A-Za-z0-9_\-]", "-", name).strip("-")
        eq_type = clean_str(r.get("Type")) or "Unknown"
        mfr = clean_str(r.get("Manufacturer"))
        data.append(
            (
                str(legacy_uuid(name)),
                code,
                name,
                eq_type,
                mfr,
                LEGACY_NOTE,
            )
        )
    if not dry:
        psycopg2.extras.execute_values(
            cur,
            """INSERT INTO equipment
                 (equipment_id, equipment_code, equipment_name, equipment_type, manufacturer, notes)
               VALUES %s
               ON CONFLICT (equipment_code) DO UPDATE SET
                 manufacturer = COALESCE(EXCLUDED.manufacturer, equipment.manufacturer)""",
            data,
        )
    return len(data)


def load_tools(cur, rows: list[dict], dry: bool) -> int:
    data = []
    for r in rows:
        uid = clean_str(r.get("Unique ID"))
        name = clean_str(r.get("Name"))
        if not (uid and name):
            continue
        tool_type = clean_str(r.get("Operation")) or clean_str(r.get("Op Type"))
        mfr = clean_str(r.get("Manufacturer"))
        datasheet = clean_str(r.get("Datasheet")) or clean_str(r.get("Link"))
        op_type = clean_str(r.get("Op Type")) or clean_str(r.get("Operation Type"))
        data.append(
            (
                str(legacy_uuid(uid)),
                uid,
                name,
                tool_type,
                mfr,
                datasheet,
                op_type,
                clean_float(
                    r.get("Cutter Diameter [mm]") or r.get("Cutter Diameter (mm)")
                ),
                clean_float(r.get("Shank Width [mm]") or r.get("Shank Width (mm)")),
                clean_float(r.get("Shank Length [mm]") or r.get("Shank Length (mm)")),
                clean_float(
                    r.get("Overall Length [mm]") or r.get("Overall Length (mm)")
                ),
                clean_str(r.get("Shank Type")),
                clean_str(r.get("Cutting Direction")),
                clean_str(r.get("Insert Clamping System") or r.get("Clamping System")),
                LEGACY_NOTE,
            )
        )
    if not dry:
        psycopg2.extras.execute_values(
            cur,
            """INSERT INTO tools
                 (tool_id, tool_code, tool_name, tool_type,
                  manufacturer, datasheet_url, op_type,
                  cutter_diameter_mm, shank_width_mm, shank_length_mm, overall_length_mm,
                  shank_type, cutting_direction, insert_clamping_system, notes)
               VALUES %s
               ON CONFLICT (tool_code) DO UPDATE SET
                 manufacturer          = COALESCE(EXCLUDED.manufacturer, tools.manufacturer),
                 datasheet_url         = COALESCE(EXCLUDED.datasheet_url, tools.datasheet_url),
                 op_type               = COALESCE(EXCLUDED.op_type, tools.op_type),
                 cutter_diameter_mm    = COALESCE(EXCLUDED.cutter_diameter_mm, tools.cutter_diameter_mm),
                 shank_width_mm        = COALESCE(EXCLUDED.shank_width_mm, tools.shank_width_mm),
                 shank_length_mm       = COALESCE(EXCLUDED.shank_length_mm, tools.shank_length_mm),
                 overall_length_mm     = COALESCE(EXCLUDED.overall_length_mm, tools.overall_length_mm),
                 shank_type            = COALESCE(EXCLUDED.shank_type, tools.shank_type),
                 cutting_direction     = COALESCE(EXCLUDED.cutting_direction, tools.cutting_direction),
                 insert_clamping_system = COALESCE(EXCLUDED.insert_clamping_system, tools.insert_clamping_system)""",
            data,
        )
    return len(data)


def load_insert_types(cur, rows: list[dict], dry: bool) -> int:
    data = []
    for r in rows:
        name = clean_str(r.get("Name"))
        if not name:
            continue
        mfr = clean_str(r.get("Manufacturer"))
        # OP Type is a Ref to OP Types table — store the display name as text
        op_type = clean_str(r.get("OP Type")) or clean_str(r.get("Op Type"))
        # Mounting style code may come as a number or short text
        msc_raw = r.get("Insert mounting Style Code (IFS)") or r.get(
            "Mounting Style Code"
        )
        mounting_style = clean_str(msc_raw)
        per_box_raw = r.get("# Per Box") or r.get("Per Box")
        edge_raw = r.get("Edge #") or r.get("Edge Number")
        # Material classification: may be an EnumList (comma-separated); take first
        mat_cls_raw = clean_str(
            r.get("Material classification level 1 (TMC1ISO)")
            or r.get("Material Classification (TMC1ISO)")
        )
        mat_cls = mat_cls_raw.split(",")[0].strip() if mat_cls_raw else None
        data.append(
            (
                str(legacy_uuid(name)),
                name,
                mfr,
                LEGACY_NOTE,
                op_type,
                mounting_style,
                int(float(per_box_raw)) if per_box_raw is not None else None,
                int(float(edge_raw)) if edge_raw is not None else None,
                clean_float(
                    r.get("Nose Radius (RE) [mm]") or r.get("Nose Radius (RE)(mm)")
                ),
                clean_float(
                    r.get("Cutting edge length (L) [mm]")
                    or r.get("Cutting edge length (L)(mm)")
                ),
                clean_float(
                    r.get("Included angle ESPR [deg]")
                    or r.get("Included angle (ESPR)(deg)")
                ),
                clean_float(
                    r.get("Fixing Hole Diameter [mm]")
                    or r.get("Fixing Hole Diameter (mm)")
                ),
                mat_cls,
            )
        )
    if not dry:
        psycopg2.extras.execute_values(
            cur,
            """INSERT INTO insert_types
                 (insert_type_id, type_code, manufacturer, geometry_notes,
                  op_type, mounting_style_code, inserts_per_box, edge_count,
                  nose_radius_mm, cutting_edge_length_mm, included_angle_deg,
                  fixing_hole_diameter_mm, material_class)
               VALUES %s
               ON CONFLICT (type_code) DO UPDATE SET
                 manufacturer          = COALESCE(EXCLUDED.manufacturer, insert_types.manufacturer),
                 op_type               = COALESCE(EXCLUDED.op_type, insert_types.op_type),
                 mounting_style_code   = COALESCE(EXCLUDED.mounting_style_code, insert_types.mounting_style_code),
                 inserts_per_box       = COALESCE(EXCLUDED.inserts_per_box, insert_types.inserts_per_box),
                 edge_count            = COALESCE(EXCLUDED.edge_count, insert_types.edge_count),
                 nose_radius_mm        = COALESCE(EXCLUDED.nose_radius_mm, insert_types.nose_radius_mm),
                 cutting_edge_length_mm = COALESCE(EXCLUDED.cutting_edge_length_mm, insert_types.cutting_edge_length_mm),
                 included_angle_deg    = COALESCE(EXCLUDED.included_angle_deg, insert_types.included_angle_deg),
                 fixing_hole_diameter_mm = COALESCE(EXCLUDED.fixing_hole_diameter_mm, insert_types.fixing_hole_diameter_mm),
                 material_class        = COALESCE(EXCLUDED.material_class, insert_types.material_class)""",
            data,
        )
    return len(data)


def load_manufacturing_methods(
    cur, op_rows: list[dict], code_rows: list[dict], dry: bool
) -> int:
    # Seed already covers FAST/MF, Forged/MO etc.  Add any missing from sheets.
    method_map = {
        "Turning": ("MC", "CNC Turning"),
        "Milling": ("MM", "CNC Milling"),
        "FAST": ("MF", "FAST/SPS Sintering"),
        "Forged": ("MO", "Forging"),
        "Rolled": ("MR", "Rolling"),
        "Cast": ("MCA", "Casting"),
        "Additive": ("MAM", "Additive Manufacturing"),
        "HIP": ("MHIP", "Hot Isostatic Pressing"),
        "EDM": ("MEDM", "Electrical Discharge Machining"),
        "Machined": ("MC2", "Machining (General)"),
        "Unknown": ("MUK", "Unknown Manufacturing Route"),
        "Other": ("MOTH", "Other Manufacturing Route"),
    }
    # Augment from Manufacturing Codes sheet
    for r in code_rows:
        meth = clean_str(r.get("Method"))
        code = clean_str(r.get("Code"))
        if meth and code and meth not in method_map:
            method_map[meth] = (code, meth)

    data = []
    for method_name, (code, display) in method_map.items():
        data.append(
            (
                str(legacy_uuid(f"method:{code}")),
                code,
                display,
                LEGACY_NOTE,
            )
        )
    if not dry:
        psycopg2.extras.execute_values(
            cur,
            """INSERT INTO manufacturing_methods
                 (method_id, method_code, method_name, description)
               VALUES %s ON CONFLICT (method_code) DO NOTHING""",
            data,
        )
    return len(data)


def load_method_parameters(cur, dry: bool) -> int:
    """No-op: the method_parameters template catalog was dropped (migration 039).
    Method parameters now live as typed inline columns on manufacturing_operations
    (see 20260623000032_inline_param_fields.sql), so there is nothing to seed."""
    return 0


def load_tool_boxes(
    cur, rows: list[dict], insert_rows: list[dict], itype_map: dict, dry: bool
) -> tuple[int, dict]:
    """Returns (count, box_id_to_uuid_map)."""
    # Find insert types by name
    if not dry:
        cur.execute("SELECT insert_type_id, type_code FROM insert_types")
        itype_uuid = {r[1]: r[0] for r in cur.fetchall()}
    else:
        itype_uuid = {}

    # Collect all box IDs referenced by inserts (to catch orphans)
    known_box_ids = {str(r.get("Box ID")) for r in rows if r.get("Box ID")}
    ref_box_ids = {str(r.get("Insert Box")) for r in insert_rows if r.get("Insert Box")}
    orphan_ids = ref_box_ids - known_box_ids

    box_uuid_map: dict[str, str] = {}
    data = []

    for r in rows:
        bid = clean_str(r.get("Box ID"))
        if not bid:
            continue
        itype_name = clean_str(r.get("Insert Type"))
        itype_id = itype_uuid.get(itype_name) if itype_name else None
        box_code = bid  # Use raw Box ID as code (unique per row)
        box_uuid = str(legacy_uuid(f"box:{bid}"))
        box_uuid_map[bid] = box_uuid
        pkg_qty_raw = r.get("Package Quantity") or r.get("Qty")
        pkg_qty = int(float(pkg_qty_raw)) if pkg_qty_raw is not None else None
        owner = clean_str(r.get("Owner"))
        data.append(
            (
                box_uuid,
                box_code,
                itype_name or "Unknown",  # description = insert type name
                clean_str(r.get("Location")),
                itype_id,
                pkg_qty,
                owner,
                LEGACY_NOTE,
            )
        )

    # Synthetic placeholder for orphan boxes referenced by inserts but missing from inventory
    for oid in orphan_ids:
        box_uuid = str(legacy_uuid(f"box:{oid}"))
        box_uuid_map[oid] = box_uuid
        data.append(
            (
                box_uuid,
                oid,
                "Unknown (not in inventory)",
                None,
                None,
                None,
                None,
                f"{LEGACY_NOTE} WARNING: box referenced by inserts but missing from Insert Boxes Inventory.",
            )
        )

    if not dry:
        psycopg2.extras.execute_values(
            cur,
            """INSERT INTO tool_boxes
                 (tool_box_id, tool_box_code, description, location,
                  insert_type_id, package_quantity, owner, notes)
               VALUES %s
               ON CONFLICT (tool_box_code) DO UPDATE SET
                 package_quantity = COALESCE(EXCLUDED.package_quantity, tool_boxes.package_quantity),
                 owner            = COALESCE(EXCLUDED.owner, tool_boxes.owner)""",
            data,
        )
    return len(data), box_uuid_map


def load_cutting_inserts(
    cur, rows: list[dict], box_uuid_map: dict, itype_map: dict, dry: bool
) -> tuple[int, dict]:
    """Returns (count, insert_uid_to_uuid_map)."""
    if not dry:
        cur.execute("SELECT insert_type_id, type_code FROM insert_types")
        itype_uuid = {r[1]: r[0] for r in cur.fetchall()}
    else:
        itype_uuid = {}

    insert_uuid_map: dict[str, str] = {}
    data = []
    seen_codes: set[str] = set()

    for r in rows:
        uid = clean_str(r.get("Unique ID"))
        if not uid:
            continue
        box_hex = clean_str(r.get("Insert Box"))
        box_uuid = box_uuid_map.get(box_hex) if box_hex else None
        if not box_uuid:
            log.warning(
                "Cutting insert %s: parent box %s not found, skipping", uid, box_hex
            )
            continue
        itype_name = clean_str(r.get("Insert Type"))
        itype_id = itype_uuid.get(itype_name) if itype_name else None
        pos = r.get("Position In Box")
        insert_number = int(float(pos)) if pos is not None else None
        # insert_code: box_code-#position
        insert_code = (
            f"{box_hex}-#{insert_number}" if insert_number else f"{box_hex}-{uid[:4]}"
        )
        if insert_code in seen_codes:
            insert_code = f"{insert_code}-{uid[:4]}"
        seen_codes.add(insert_code)
        status = clean_str(r.get("Status"))
        is_depleted = status and status.lower() in ("used", "depleted", "consumed")
        location = clean_str(r.get("Location"))
        owner = clean_str(r.get("Owner"))
        insert_uuid = str(legacy_uuid(f"insert:{uid}"))
        insert_uuid_map[uid] = insert_uuid
        data.append(
            (
                insert_uuid,
                insert_code,
                box_uuid,
                itype_id,
                insert_number,
                location,
                owner,
                is_depleted,
                LEGACY_NOTE,
            )
        )

    if not dry:
        psycopg2.extras.execute_values(
            cur,
            """INSERT INTO cutting_inserts
                 (insert_id, insert_code, tool_box_id, insert_type_id,
                  insert_number, location, owner, is_depleted, notes)
               VALUES %s
               ON CONFLICT (insert_code) DO UPDATE SET
                 location = COALESCE(EXCLUDED.location, cutting_inserts.location),
                 owner    = COALESCE(EXCLUDED.owner, cutting_inserts.owner)""",
            data,
        )
    return len(data), insert_uuid_map


def load_insert_edges(
    cur, rows: list[dict], insert_uuid_map: dict, box_uuid_map: dict, dry: bool
) -> tuple[int, dict]:
    """Returns (count, edge_uid_to_uuid_map).

    edge_code = {box_hex}-#{insert_position}-f{letter}
    where letter = A/B/C/D based on edge Number (1→A, 2→B, ...).
    """
    edge_uuid_map: dict[str, str] = {}
    data = []
    seen_codes: set[str] = set()

    # Build a map: insert_uid → (box_hex, insert_number) for code derivation
    if not dry:
        cur.execute(
            "SELECT ci.insert_code, ci.insert_number, tb.tool_box_code "
            "FROM cutting_inserts ci JOIN tool_boxes tb ON ci.tool_box_id=tb.tool_box_id"
        )
        insert_info: dict[str, tuple] = {r[0]: (r[2], r[1]) for r in cur.fetchall()}  # noqa: F841
    else:
        insert_info = {}  # noqa: F841

    for r in rows:
        eid = clean_str(r.get("Edge ID"))
        if not eid:
            continue
        parent_uid = clean_str(r.get("Parent Insert"))
        insert_uuid = insert_uuid_map.get(parent_uid) if parent_uid else None
        if not insert_uuid:
            log.warning(
                "Insert edge %s: parent insert %s not found, skipping", eid, parent_uid
            )
            continue

        number = r.get("Number")
        edge_num = int(float(number)) if number is not None else 1
        # edge letter: A=1, B=2, C=3, D=4, ...
        edge_letter = chr(ord("A") + edge_num - 1) if edge_num <= 26 else str(edge_num)
        edge_identifier = f"f{edge_letter}"

        parent_box_hex = clean_str(r.get("Parent Box"))
        # Look up box code + insert number from DB (or derive from map)
        box_code = parent_box_hex or "UNK"
        # find insert number: look at insert_uuid_map to get the insert code
        # which was: f"{box_hex}-#{insert_number}"
        insert_number = None  # noqa: F841
        for k, v in insert_uuid_map.items():
            if v == insert_uuid:
                # k is the original uid; look at data to get position
                pass
        # Simpler: just use edge_identifier as unique suffix within insert
        edge_code = f"{box_code}-{parent_uid[:4]}-{edge_identifier}"
        if edge_code in seen_codes:
            edge_code = f"{edge_code}-{eid[:4]}"
        seen_codes.add(edge_code)

        is_used = clean_bool(r.get("Status"))
        edge_uuid = str(legacy_uuid(f"edge:{eid}"))
        edge_uuid_map[eid] = edge_uuid
        data.append(
            (
                edge_uuid,
                edge_code,
                insert_uuid,
                edge_identifier,
                is_used,
                LEGACY_NOTE,
            )
        )

    if not dry:
        psycopg2.extras.execute_values(
            cur,
            """INSERT INTO insert_edges
                 (edge_id, edge_code, insert_id, edge_identifier, is_used, notes)
               VALUES %s ON CONFLICT (edge_code) DO NOTHING""",
            data,
        )
    return len(data), edge_uuid_map


def load_physical_samples(
    cur, rows: list[dict], material_map: dict, email_to_uuid_map: dict, dry: bool
) -> tuple[int, dict, dict]:
    """Returns (count, legacy_uid_to_uuid_map, sample_code_to_uuid_map).

    Imports all Inventory rows (both 'Sample' and 'Equipment' item types).
    Equipment items are mapped to form='Equipment' and noted accordingly.
    sample_code_map is keyed by Item Code (human-readable sample code) so that
    load_machining_ops can resolve Workpiece IDs stored as sample codes.
    """
    sample_uuid_map: dict[str, str] = {}
    sample_code_map: dict[str, str] = {}
    data = []

    for r in rows:
        uid = clean_str(r.get("Unique ID"))
        code = clean_str(r.get("Item Code"))
        if not (uid and code):
            continue
        alloy_name = clean_str(r.get("Alloy"))
        mat_id = material_map.get(alloy_name)
        # Geometry = physical shape; Item Type = category (sample/equipment/misc)
        geo = clean_str(r.get("Geometry")) or "other"
        item_type_raw = clean_str(r.get("Item Type"))
        item_type = item_type_raw.lower() if item_type_raw else "sample"
        # Normalise item_type to allowed CHECK values
        if item_type not in ("sample", "equipment", "miscellaneous"):
            item_type = "sample"
        mfg_date = clean_date(r.get("Manufacturing Date"))
        export_ctrl = clean_bool(r.get("Export Controlled?"))
        extra_notes_parts = []
        if alloy_name and not mat_id:
            extra_notes_parts.append(f"Legacy alloy (unresolved): {alloy_name}")
        notes = (
            " | ".join([LEGACY_NOTE] + extra_notes_parts)
            if extra_notes_parts
            else LEGACY_NOTE
        )

        # Manufacturing Route: stored as a Ref display name in AppSheet
        mfg_route = clean_str(r.get("Manufacturing Route"))

        # Co-owners: EnumList in AppSheet — may be comma-separated emails/names
        co_owners_raw = r.get("Co-owners") or r.get("Co-Owners")
        co_owners = clean_str(co_owners_raw)

        s_uuid = str(legacy_uuid(f"sample:{uid}"))
        sample_uuid_map[uid] = s_uuid
        if code:
            sample_code_map[code] = s_uuid
        data.append(
            (
                s_uuid,
                code,
                mat_id,
                geo.lower(),
                item_type,
                clean_float(r.get("Weight [g]")),
                clean_float(r.get("⌀ [mm]")),
                clean_float(r.get("x [mm]")),  # x = width
                clean_float(r.get("z [mm]")),
                clean_float(r.get("y [mm]")),
                "active",
                mfg_date,
                export_ctrl,
                notes,
                clean_str(r.get("Nickname")),
                clean_str(r.get("Location")),
                clean_str(r.get("Surface Finish")),
                email_to_uuid_map.get((clean_str(r.get("Owner")) or "").lower()),
                co_owners,
                mfg_route,
                clean_bool(r.get("Mounted")),
                clean_str(r.get("Mounting_Method") or r.get("Mounting Method")),
                clean_str(r.get("Notes")),
            )
        )

    if not dry:
        psycopg2.extras.execute_values(
            cur,
            """INSERT INTO physical_samples
                 (sample_id, sample_code, material_id, form, item_type, mass_grams,
                  diameter_mm, width_mm, length_mm, thickness_mm, current_status,
                  manufactured_date, export_controlled, notes,
                  nickname, location, surface_finish,
                  owner, co_owners, manufacturing_route, mounted, mounting_method,
                  legacy_notes)
               VALUES %s
               ON CONFLICT (sample_code) DO UPDATE SET
                 form                = EXCLUDED.form,
                 item_type           = COALESCE(EXCLUDED.item_type, physical_samples.item_type),
                 nickname            = COALESCE(EXCLUDED.nickname, physical_samples.nickname),
                 location            = COALESCE(EXCLUDED.location, physical_samples.location),
                 surface_finish      = COALESCE(EXCLUDED.surface_finish, physical_samples.surface_finish),
                 owner               = COALESCE(EXCLUDED.owner, physical_samples.owner),
                 co_owners           = COALESCE(EXCLUDED.co_owners, physical_samples.co_owners),
                 manufacturing_route = COALESCE(EXCLUDED.manufacturing_route, physical_samples.manufacturing_route),
                 mounted             = COALESCE(EXCLUDED.mounted, physical_samples.mounted),
                 mounting_method     = COALESCE(EXCLUDED.mounting_method, physical_samples.mounting_method),
                 width_mm            = COALESCE(EXCLUDED.width_mm, physical_samples.width_mm),
                 legacy_notes        = COALESCE(EXCLUDED.legacy_notes, physical_samples.legacy_notes),
                 notes               = EXCLUDED.notes""",
            data,
        )
    return len(data), sample_uuid_map, sample_code_map


def load_users(cur, rows: list[dict], dry: bool) -> dict[str, str]:
    """Upsert Directus users from the XLSX Users sheet.

    Returns {email.lower(): uuid_str} for resolving owner/co-owner references.
    Carolina Guerra is seeded unconditionally — she appears in co_owners data
    but is absent from the Users sheet.
    """

    def _user_uuid(email: str) -> str:
        return str(legacy_uuid(f"d1_user:{email.lower()}"))

    email_to_uuid: dict[str, str] = {}
    rows_to_insert = []

    for r in rows:
        first = clean_str(
            r.get("First Name") or r.get("FirstName") or r.get("first_name")
        )
        last = clean_str(
            r.get("Last Name")
            or r.get("LastName")
            or r.get("Surname")
            or r.get("last_name")
        )
        email = clean_str(r.get("Email") or r.get("Email Address") or r.get("email"))
        role_name = clean_str(r.get("Role") or r.get("role"))
        if not email:
            continue
        email_lc = email.lower()
        u_uuid = _user_uuid(email_lc)
        email_to_uuid[email_lc] = u_uuid
        role_id = (
            _LAB_ADMIN_ROLE_ID
            if role_name and "admin" in role_name.lower()
            else _LAB_MEMBER_ROLE_ID
        )
        rows_to_insert.append(
            (u_uuid, first or "Unknown", last or "Unknown", email_lc, role_id)
        )

    # Carolina Guerra: found in co_owners but not in Users sheet
    _cg_email = "carolina.guerra@nottingham.ac.uk"
    if _cg_email not in email_to_uuid:
        cg_uuid = _user_uuid(_cg_email)
        email_to_uuid[_cg_email] = cg_uuid
        rows_to_insert.append(
            (cg_uuid, "Carolina", "Guerra", _cg_email, _LAB_MEMBER_ROLE_ID)
        )

    if not dry and rows_to_insert:
        psycopg2.extras.execute_values(
            cur,
            """INSERT INTO directus_users
                 (id, first_name, last_name, email, role, status, password)
               VALUES %s
               ON CONFLICT (email) DO UPDATE SET
                 first_name = COALESCE(directus_users.first_name, EXCLUDED.first_name),
                 last_name  = COALESCE(directus_users.last_name,  EXCLUDED.last_name),
                 role       = COALESCE(directus_users.role,       EXCLUDED.role)""",
            [(d[0], d[1], d[2], d[3], d[4], "invited", None) for d in rows_to_insert],
        )

    return email_to_uuid


def load_co_owners(cur, email_to_uuid_map: dict[str, str], dry: bool) -> int:
    """Migrate co_owners TEXT column → sample_co_owners junction rows.

    Reads every physical_samples row where co_owners IS NOT NULL, splits on
    common delimiters, resolves each email to a UUID, and inserts junction rows.
    Returns the number of rows inserted.
    """
    if dry:
        return 0

    cur.execute(
        "SELECT sample_id, co_owners FROM physical_samples WHERE co_owners IS NOT NULL"
    )
    db_rows = cur.fetchall()
    data = []
    for sample_id, co_owners_text in db_rows:
        parts = re.split(r"[;,\s]+", (co_owners_text or "").strip())
        for part in parts:
            email = part.strip().lower()
            if not email or "@" not in email:
                continue
            user_uuid = email_to_uuid_map.get(email)
            if not user_uuid:
                log.warning(
                    "co_owners: no UUID for %s (sample %s), skipping", email, sample_id
                )
                continue
            data.append((str(uuid.uuid4()), str(sample_id), user_uuid))

    if data:
        psycopg2.extras.execute_values(
            cur,
            """INSERT INTO sample_co_owners (id, sample_id, user_id) VALUES %s
               ON CONFLICT (sample_id, user_id) DO NOTHING""",
            data,
        )
    return len(data)


def load_sample_genealogy(
    cur, rows: list[dict], sample_uuid_map: dict, dry: bool
) -> tuple[int, list[str]]:
    """Insert parent→child genealogy from Inventory[Parent] field.

    Returns (inserted, [skipped_reasons]).
    """
    data = []
    skipped = []

    for r in rows:
        child_uid = clean_str(r.get("Unique ID"))
        parent_uid = clean_str(r.get("Parent"))
        if not (child_uid and parent_uid):
            continue
        child_uuid = sample_uuid_map.get(child_uid)
        parent_uuid = sample_uuid_map.get(parent_uid)
        if not child_uuid:
            skipped.append(f"child {child_uid} not in sample_uuid_map")
            continue
        if not parent_uuid:
            skipped.append(
                f"parent {parent_uid} for child {r.get('Item Code')} not found"
            )
            continue
        if child_uuid == parent_uuid:
            skipped.append(f"self-loop on {child_uid}")
            continue
        data.append((child_uuid, parent_uuid, "cut_from"))

    # Deduplicate
    data = list({(c, p, rt) for c, p, rt in data})

    if not dry:
        psycopg2.extras.execute_values(
            cur,
            """INSERT INTO sample_genealogy
                 (child_sample_id, parent_sample_id, relationship_type)
               VALUES %s ON CONFLICT DO NOTHING""",
            data,
        )
    return len(data), skipped


def load_fast_runs(
    cur,
    fast_rows: list[dict],
    inv_rows: list[dict],
    sample_uuid_map: dict,
    sample_code_map: dict,
    equipment_map: dict,
    method_map: dict,
    dry: bool,
) -> tuple[int, int]:
    """Import FAST Runs and back-fill sintering_params. Returns (imported, skipped_no_sample)."""
    # Build reverse map: fast_run_uid → list of sample uids
    fast_to_samples: dict[str, list[str]] = defaultdict(list)
    for r in inv_rows:
        sid = clean_str(r.get("Unique ID"))
        mfg_op = clean_str(r.get("Manufacturing Operation ID"))
        if sid and mfg_op:
            fast_to_samples[mfg_op].append(sid)

    fast_method_id = method_map.get("MF")
    data = []
    sintering_data: list[tuple] = []
    skipped_no_sample = 0

    for r in fast_rows:
        uid = clean_str(r.get("Unique ID"))
        if not uid:
            continue
        sample_uids = fast_to_samples.get(uid, [])
        if not sample_uids:
            skipped_no_sample += 1
            continue

        op_date = clean_date(r.get("Date")) or clean_date(r.get("Process DateTime"))
        machine_name = clean_str(r.get("Machine"))
        equipment_id = equipment_map.get(machine_name)
        operator = clean_str(r.get("User"))
        notes_text = clean_str(r.get("Notes"))

        metadata: dict = {}
        for key, col in [
            ("recipe_number", "Recipe #"),
            ("batch_number", "Batch #"),
            ("atmosphere", "Atmosphere"),
            ("tc_pyro_control", "TC/Pyro control"),
            ("material_type", "Material Type"),
            ("coshh_ref", "CoSHH Ref #"),
        ]:
            v = clean_str(r.get(col))
            if v:
                metadata[key] = v
        for key, col in [
            ("mass_grams", "Mass (g)"),
            ("mould_diameter_mm", "Mould diameter (mm)"),
            ("max_force_kn", "Max Force (kN)"),
            ("max_temp_celsius", "Max Temp (°C)"),
            ("voltage_at_max_t_v", "Voltage at Max T (V)"),
            ("power_at_max_t_kw", "Power at Max T (kW)"),
            ("ptc_top_celsius", "PTC top (°C)"),
            ("ptc_bot_celsius", "PTC bot (°C)"),
        ]:
            v = clean_float(r.get(col))
            if v is not None:
                metadata[key] = v
        material_free_text = clean_str(r.get("Material"))
        if material_free_text:
            metadata["legacy_material_text"] = material_free_text
        metadata["legacy_fast_run_uid"] = uid

        outcome = notes_text or ""
        outcome = f"{LEGACY_NOTE} {outcome}".strip()

        material_note = metadata.get("material_type") or metadata.get(
            "legacy_material_text"
        )

        for sample_uid in sample_uids:
            s_uuid = sample_uuid_map.get(sample_uid) or sample_code_map.get(sample_uid)
            if not s_uuid:
                continue
            op_uuid = str(legacy_uuid(f"fast_op:{uid}:{sample_uid}"))
            data.append(
                (
                    op_uuid,
                    s_uuid,
                    fast_method_id,
                    equipment_id,
                    operator,
                    op_date,
                    json.dumps(metadata),
                    outcome,
                )
            )
            sintering_data.append(
                (
                    op_uuid,
                    metadata.get("recipe_number"),
                    metadata.get("batch_number"),
                    metadata.get("mould_diameter_mm"),
                    metadata.get("atmosphere"),
                    metadata.get("tc_pyro_control"),
                    metadata.get("max_temp_celsius"),
                    metadata.get("max_force_kn"),
                    metadata.get("voltage_at_max_t_v"),
                    metadata.get("power_at_max_t_kw"),
                    metadata.get("ptc_top_celsius"),
                    metadata.get("ptc_bot_celsius"),
                    metadata.get("coshh_ref"),
                    material_note,
                )
            )

    if not dry:
        psycopg2.extras.execute_values(
            cur,
            """INSERT INTO manufacturing_operations
                 (operation_id, sample_id, method_id, equipment_id,
                  operator_name, operation_date, recorded_metadata, outcome_notes)
               VALUES %s ON CONFLICT DO NOTHING""",
            data,
        )
        if sintering_data:
            psycopg2.extras.execute_values(
                cur,
                """INSERT INTO sintering_params
                     (operation_id, recipe_number, batch_number,
                      mould_diameter_mm, atmosphere, tc_pyro_control,
                      max_temp_celsius, max_force_kn,
                      voltage_at_max_t_v, power_at_max_t_kw,
                      ptc_top_celsius, ptc_bot_celsius,
                      coshh_ref, material_type_note)
                   VALUES %s ON CONFLICT (operation_id) DO NOTHING""",
                sintering_data,
            )
    return len(data), skipped_no_sample


_OP_SUBTYPE_MAP = {
    "turning": "turning",
    "facing": "facing",
    "boring": "boring",
    "threading": "threading",
    "grooving": "grooving",
    "parting": "parting",
    "milling": "milling",
    "drilling": "drilling",
}


def load_machining_ops(
    cur,
    rows: list[dict],
    sample_uuid_map: dict,
    sample_code_map: dict,
    equipment_map: dict,
    tool_uuid_map: dict,
    method_map: dict,
    dry: bool,
) -> tuple[int, list[str]]:
    """Import Machining Operations and back-fill machining_params. Returns (imported, [warnings])."""
    warnings_out: list[str] = []
    data = []
    param_data = []

    for r in rows:
        uid = clean_str(r.get("Unique ID"))
        if not uid:
            continue
        workpiece_uid = clean_str(r.get("Workpiece ID"))
        s_uuid = (
            (sample_uuid_map.get(workpiece_uid) or sample_code_map.get(workpiece_uid))
            if workpiece_uid
            else None
        )
        if not s_uuid:
            warnings_out.append(
                f"Machining op {uid}: workpiece {workpiece_uid} not found — skipped"
            )
            continue

        operation_type = clean_str(r.get("Operation")) or "Turning"
        method_code = "MM" if operation_type.lower() == "milling" else "MC"
        method_id = method_map.get(method_code)

        machine_name = clean_str(r.get("Machine"))
        equipment_id = equipment_map.get(machine_name)

        # Tool: "(hex_id) Name" format
        tool_raw = clean_str(r.get("Tool"))
        tool_hex = extract_tool_uuid(tool_raw) if tool_raw else None
        tool_uuid = tool_uuid_map.get(tool_hex) if tool_hex else None

        # insert_edge_id: legacy human-readable codes; store in metadata only
        legacy_edge_id = clean_str(r.get("Insert Edge ID"))

        abs_pass = r.get("Abs Pass #")
        op_seq = None
        if abs_pass is not None and str(abs_pass).strip():
            try:
                op_seq = int(float(abs_pass))
            except (ValueError, TypeError):
                pass

        freq_hz = clean_float(r.get("Capture Frequency [Hz]"))
        freq_khz = freq_hz / 1000.0 if freq_hz else None

        sw = clean_str(r.get("Software Used"))
        ver = clean_str(r.get("Version"))
        capture_sw = " ".join(filter(None, [sw, ver])) or None

        force_link = clean_str(r.get("Force File Link"))
        nc_prog = clean_str(r.get("NC Program"))
        op_date_raw = r.get("Creation Date")
        op_date = clean_date(op_date_raw)

        metadata: dict = {"legacy_machining_uid": uid}
        for key, col in [
            ("op_type", "Op Type"),
            ("coolant_pressure", "Coolant Pressure"),
            ("chips_ref_code", "Chips Ref Code"),
            ("experiment_sheet", "Experiment Sheet"),
            ("operation_code", "Operation Code"),
        ]:
            v = clean_str(r.get(col))
            if v:
                metadata[key] = v
        for key, col in [
            ("rpm", "RPM"),
            ("vc_m_per_min", "Vc [m/min]"),
            ("feed_mm_per_rev", "FR [mm/rev]"),
            ("axial_mm", "Axial [mm]"),
            ("diameter_mm", "⌀ [mm]"),
            ("ap_mm", "Ap [mm]"),
            ("cut_length_mm", "Cut Length [mm]"),
            ("max_rpm", "Max RPM"),
        ]:
            v = clean_float(r.get(col))
            if v is not None:
                metadata[key] = v
        for key, col in [
            ("force_captured", "Force Captured"),
            ("new_edge", "New Edge?"),
            ("tacho_used", "Tacho Used"),
            ("coolant_used", "Coolant Used"),
            ("chips_collected", "Chips Collected"),
        ]:
            raw = r.get(col)
            if raw is not None:
                metadata[key] = clean_bool(raw)
        if legacy_edge_id and legacy_edge_id.upper() != "N/A":
            metadata["legacy_insert_edge_id"] = legacy_edge_id

        notes = clean_str(r.get("Notes")) or ""
        outcome = f"{LEGACY_NOTE} {notes}".strip()
        pass_code = clean_str(r.get("Operation Code"))

        op_uuid = str(legacy_uuid(f"machining_op:{uid}"))
        data.append(
            (
                op_uuid,
                s_uuid,
                method_id,
                equipment_id,
                tool_uuid,
                clean_str(r.get("User")),
                op_seq,
                pass_code,
                op_date,
                json.dumps(metadata),
                capture_sw,
                freq_khz,
                force_link,
                nc_prog,
                outcome,
            )
        )

        # Back-fill machining_params from extracted metadata
        subtype_raw = operation_type.lower().strip()
        subtype = _OP_SUBTYPE_MAP.get(subtype_raw, "other")
        rpm_val = metadata.get("rpm") or metadata.get("max_rpm")
        axial_val = metadata.get("axial_mm") or metadata.get("ap_mm")
        coolant_p_raw = metadata.get("coolant_pressure")
        coolant_p = (
            clean_float(str(coolant_p_raw)) if coolant_p_raw is not None else None
        )
        param_data.append(
            (
                op_uuid,
                subtype,
                rpm_val,
                metadata.get("vc_m_per_min"),
                metadata.get("feed_mm_per_rev"),
                axial_val,
                None,  # radial_depth_of_cut_mm — not in legacy data
                metadata.get("cut_length_mm"),
                metadata.get("diameter_mm"),
                metadata.get("new_edge"),
                metadata.get("coolant_used"),
                coolant_p,
                metadata.get("tacho_used"),
                metadata.get("force_captured"),
                metadata.get("chips_collected"),
                metadata.get("chips_ref_code"),
                metadata.get("experiment_sheet"),
                metadata.get("legacy_insert_edge_id"),
                uid,  # legacy_machining_uid
            )
        )

    if not dry:
        psycopg2.extras.execute_values(
            cur,
            """INSERT INTO manufacturing_operations
                 (operation_id, sample_id, method_id, equipment_id, tool_id,
                  operator_name, operation_sequence, pass_code, operation_date,
                  recorded_metadata, capture_software, capture_frequency_khz,
                  file_storage_pointer, nc_program_text, outcome_notes)
               VALUES %s ON CONFLICT DO NOTHING""",
            data,
        )
        if param_data:
            psycopg2.extras.execute_values(
                cur,
                """INSERT INTO machining_params
                     (operation_id, operation_subtype,
                      spindle_speed_rpm, cutting_speed_m_per_min, feed_mm_per_rev,
                      axial_depth_of_cut_mm, radial_depth_of_cut_mm, cutting_length_mm,
                      workpiece_diameter_mm,
                      new_edge, coolant_used, coolant_pressure_bar,
                      tacho_used, force_captured, chips_collected,
                      chips_ref_code, experiment_sheet_url,
                      legacy_insert_edge_id, legacy_machining_uid)
                   VALUES %s ON CONFLICT (operation_id) DO NOTHING""",
                param_data,
            )
    return len(data), warnings_out


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def build_lookup_maps(cur, dry: bool) -> tuple[dict, dict, dict, dict]:
    """Return (material_map, equipment_map, tool_map, method_map) from DB."""
    if dry:
        return {}, {}, {}, {}
    cur.execute("SELECT common_name, material_id FROM materials")
    material_map = {r[0]: r[1] for r in cur.fetchall()}
    # Also map by alloy_code for FAST run alloy lookups
    cur.execute("SELECT alloy_code, material_id FROM materials")
    material_map.update({r[0]: r[1] for r in cur.fetchall()})

    cur.execute("SELECT equipment_name, equipment_id FROM equipment")
    equipment_map = {r[0]: r[1] for r in cur.fetchall()}

    cur.execute("SELECT tool_code, tool_id FROM tools")
    tool_map = {r[0]: r[1] for r in cur.fetchall()}

    cur.execute("SELECT method_code, method_id FROM manufacturing_methods")
    method_map = {r[0]: r[1] for r in cur.fetchall()}

    return material_map, equipment_map, tool_map, method_map


def print_report(stats: dict, warnings: list[str]) -> None:
    print("\n" + "=" * 60)
    print("D1 Legacy Migration — Reconciliation Report")
    print("=" * 60)
    print(f"\n{'Table':<40} {'Rows processed':>15}")
    print("-" * 56)
    for table, count in stats.items():
        print(f"  {table:<38} {count:>15}")
    total = sum(stats.values())
    print("-" * 56)
    print(f"  {'TOTAL':<38} {total:>15}")
    if warnings:
        print(f"\n{'─' * 60}")
        print(f"WARNINGS / SKIPPED ({len(warnings)}):")
        for w in warnings[:50]:
            print(f"  ⚠ {w}")
        if len(warnings) > 50:
            print(f"  ... and {len(warnings) - 50} more (see log)")
    print("\n✓ Migration complete.\n")


def main() -> None:
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--xlsx", required=True, help="Path to Sample_Data.xlsx")
    p.add_argument(
        "--dry-run", action="store_true", help="Print counts only; make no DB changes"
    )
    args = p.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        sys.exit(f"ERROR: {xlsx_path} not found")

    db_url = os.environ.get("DATABASE_URL")
    if not db_url and not args.dry_run:
        sys.exit("ERROR: DATABASE_URL env var required (or use --dry-run)")

    log.info("Loading %s …", xlsx_path.name)
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    # Load all sheets
    _, ae_rows = sheet_rows(wb, "Alloying Elements")
    _, alloy_rows = sheet_rows(wb, "Alloy Codes")
    _, machine_rows = sheet_rows(wb, "Machines")
    _, tool_rows = sheet_rows(wb, "Tools")
    _, itype_rows = sheet_rows(wb, "Insert Types")
    _, op_rows = sheet_rows(wb, "Operation")
    _, mfg_code_rows = sheet_rows(wb, "Manufacturing Codes")
    _, box_rows = sheet_rows(wb, "Insert Boxes Inventory")
    _, insert_rows = sheet_rows(wb, "Inserts")
    _, edge_rows = sheet_rows(wb, "Inserts Edges")
    _, inv_rows = sheet_rows(wb, "Inventory")
    _, fast_rows = sheet_rows(wb, "FAST Runs")
    _, mop_rows = sheet_rows(wb, "Machining Operations")
    _, user_rows = sheet_rows(wb, "Users", required=False)

    stats: dict[str, int] = {}
    all_warnings: list[str] = []

    if args.dry_run:
        log.info("DRY RUN — no database writes")
        conn = None
        cur = None
    else:
        conn = psycopg2.connect(db_url)
        conn.autocommit = False
        cur = conn.cursor()

    try:
        log.info("Loading alloying_elements …")
        stats["alloying_elements"] = load_alloying_elements(cur, ae_rows, args.dry_run)

        log.info("Loading materials …")
        stats["materials"] = load_materials(cur, alloy_rows, args.dry_run)

        log.info("Loading material_alloying_elements …")
        stats["material_alloying_elements"] = load_material_elements(
            cur, alloy_rows, args.dry_run
        )

        log.info("Loading equipment …")
        stats["equipment"] = load_equipment(cur, machine_rows, args.dry_run)

        log.info("Loading tools …")
        stats["tools"] = load_tools(cur, tool_rows, args.dry_run)

        log.info("Loading insert_types …")
        stats["insert_types"] = load_insert_types(cur, itype_rows, args.dry_run)

        log.info("Loading manufacturing_methods …")
        stats["manufacturing_methods"] = load_manufacturing_methods(
            cur, op_rows, mfg_code_rows, args.dry_run
        )

        log.info("Loading method_parameters …")
        stats["method_parameters"] = load_method_parameters(cur, args.dry_run)

        log.info("Loading tool_boxes …")
        n_boxes, box_uuid_map = load_tool_boxes(
            cur, box_rows, insert_rows, {}, args.dry_run
        )
        stats["tool_boxes"] = n_boxes

        log.info("Loading cutting_inserts …")
        n_inserts, insert_uuid_map = load_cutting_inserts(
            cur, insert_rows, box_uuid_map, {}, args.dry_run
        )
        stats["cutting_inserts"] = n_inserts

        log.info("Loading insert_edges …")
        n_edges, edge_uuid_map = load_insert_edges(
            cur, edge_rows, insert_uuid_map, box_uuid_map, args.dry_run
        )
        stats["insert_edges"] = n_edges

        # Refresh lookup maps from DB (now populated)
        material_map, equipment_map, tool_map, method_map = build_lookup_maps(
            cur, args.dry_run
        )

        log.info("Loading users …")
        email_to_uuid_map = load_users(cur, user_rows, args.dry_run)
        stats["directus_users"] = len(email_to_uuid_map)

        log.info("Loading physical_samples …")
        n_samples, sample_uuid_map, sample_code_map = load_physical_samples(
            cur, inv_rows, material_map, email_to_uuid_map, args.dry_run
        )
        stats["physical_samples"] = n_samples

        log.info("Loading sample_genealogy …")
        n_gen, gen_skip = load_sample_genealogy(
            cur, inv_rows, sample_uuid_map, args.dry_run
        )
        stats["sample_genealogy"] = n_gen
        all_warnings.extend(gen_skip)

        log.info("Loading FAST Run operations …")
        n_fast, fast_skipped = load_fast_runs(
            cur,
            fast_rows,
            inv_rows,
            sample_uuid_map,
            sample_code_map,
            equipment_map,
            method_map,
            args.dry_run,
        )
        stats["manufacturing_operations (FAST)"] = n_fast
        if fast_skipped:
            all_warnings.append(
                f"{fast_skipped} FAST Runs skipped — no linked Inventory sample "
                f"(Manufacturing Operation ID not set in Inventory sheet)"
            )

        log.info("Loading Machining Operations …")
        n_mop, mop_warn = load_machining_ops(
            cur,
            mop_rows,
            sample_uuid_map,
            sample_code_map,
            equipment_map,
            tool_map,
            method_map,
            args.dry_run,
        )
        stats["manufacturing_operations (Machining)"] = n_mop
        all_warnings.extend(mop_warn)

        log.info("Migrating co_owners → sample_co_owners junction …")
        stats["sample_co_owners"] = load_co_owners(cur, email_to_uuid_map, args.dry_run)

        if not args.dry_run:
            conn.commit()
            log.info("Transaction committed.")

    except Exception:
        if conn:
            conn.rollback()
        log.exception("Migration failed — rolled back")
        raise
    finally:
        if cur:
            cur.close()
        if conn:
            conn.close()

    print_report(stats, all_warnings)


if __name__ == "__main__":
    main()
